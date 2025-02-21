# -*- coding: utf-8 -*-
# @Time    : 2020/3/30 11:06
# @Author  : Hui Wang

import numpy as np
import math
import random
import os
import json
import pickle
from scipy.sparse import csr_matrix
from texttable import Texttable
import torch
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
import openpyxl
import torch.nn.functional as F
import pandas as pd
import ast

def set_seed(seed):
    random.seed(seed)
    os.environ['PYTHONHASHSEED'] = str(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    torch.cuda.manual_seed(seed)
    torch.cuda.manual_seed_all(seed)
    # some cudnn methods can be random even after fixing the seed
    # unless you tell it to be deterministic
    torch.backends.cudnn.deterministic = True

def check_path(path):
    if not os.path.exists(path):
        os.makedirs(path)
        print(f'{path} created')

def neg_sample(item_set, item_size):  # 前闭后闭
    item = random.randint(1, item_size - 1)
    while item in item_set:
        item = random.randint(1, item_size - 1)
    return item

class EarlyStopping:
    """Early stops the training if validation loss doesn't improve after a given patience."""
    def __init__(self, checkpoint_path, patience=7, verbose=False, delta=0):
        """
        Args:
            patience (int): How long to wait after last time validation loss improved.
                            Default: 7
            verbose (bool): If True, prints a message for each validation loss improvement.
                            Default: False
            delta (float): Minimum change in the monitored quantity to qualify as an improvement.
                            Default: 0
        """
        self.checkpoint_path = checkpoint_path
        self.patience = patience
        self.verbose = verbose
        self.counter = 0
        self.best_score = None
        self.early_stop = False
        self.delta = delta

    def compare(self, score):
        for i in range(len(score)):
            # 有一个指标增加了就认为是还在涨
            if score[i] > self.best_score[i]+self.delta:
                return False
        return True

    def __call__(self, score, model):
        # score HIT@10 NDCG@10

        if self.best_score is None:
            self.best_score = score
            self.score_min = np.array([0]*len(score))
            self.save_checkpoint(score, model)
        elif self.compare(score):
            self.counter += 1
            print(f'EarlyStopping counter: {self.counter} out of {self.patience}')
            if self.counter >= self.patience:
                self.early_stop = True
        else:
            self.best_score = score
            self.save_checkpoint(score, model)
            self.counter = 0

    def save_checkpoint(self, score, model):
        '''Saves model when validation loss decrease.'''
        if self.verbose:
            # ({self.score_min:.6f} --> {score:.6f}) # 这里如果是一个值的话输出才不会有问题
            print(f'Validation score increased.  Saving model ...')
        torch.save(model.state_dict(), self.checkpoint_path)
        self.score_min = score

def kmax_pooling(x, dim, k):
    index = x.topk(k, dim=dim)[1].sort(dim=dim)[0]
    return x.gather(dim, index).squeeze(dim)

def avg_pooling(x, dim):
    return x.sum(dim=dim)/x.size(dim)


def generate_rating_matrix_valid(user_seq, num_users, num_items):
    # three lists are used to construct sparse matrix
    row = []
    col = []
    data = []
    for user_id, item_list in enumerate(user_seq):
        for item in item_list[:-2]: #
            row.append(user_id)
            col.append(item)
            data.append(1)

    row = np.array(row)
    col = np.array(col)
    data = np.array(data)
    rating_matrix = csr_matrix((data, (row, col)), shape=(num_users, num_items))

    return rating_matrix

def generate_rating_matrix_test(user_seq, num_users, num_items):
    # three lists are used to construct sparse matrix
    row = []
    col = []
    data = []
    for user_id, item_list in enumerate(user_seq):
        for item in item_list[:-1]: #
            row.append(user_id)
            col.append(item)
            data.append(1)

    row = np.array(row)
    col = np.array(col)
    data = np.array(data)
    rating_matrix = csr_matrix((data, (row, col)), shape=(num_users, num_items))

    return rating_matrix

def get_user_seqs(data_file):
    lines = open(data_file).readlines()
    user_seq = []
    item_set = set()
    for line in lines:
        user, items = line.strip().split(' ', 1)
        items = items.split(' ')
        items = [int(item) for item in items]
        user_seq.append(items)
        item_set = item_set | set(items)
    max_item = max(item_set)

    num_users = len(lines)
    num_items = max_item + 2

    valid_rating_matrix = generate_rating_matrix_valid(user_seq, num_users, num_items)
    test_rating_matrix = generate_rating_matrix_test(user_seq, num_users, num_items)
    return user_seq, max_item, valid_rating_matrix, test_rating_matrix

def get_user_seqs_long(data_file):
    lines = open(data_file).readlines()
    user_seq = []
    long_sequence = []
    item_set = set()
    for line in lines:
        user, items = line.strip().split(' ', 1)
        items = items.split(' ')
        items = [int(item) for item in items]
        long_sequence.extend(items) # 后面的都是采的负例
        user_seq.append(items)
        item_set = item_set | set(items)
    max_item = max(item_set)

    return user_seq, max_item, long_sequence

def get_user_seqs_and_sample(data_file, sample_file):
    lines = open(data_file).readlines()
    user_seq = []
    item_set = set()
    for line in lines:
        user, items = line.strip().split(' ', 1)
        items = items.split(' ')
        items = [int(item) for item in items]
        user_seq.append(items)
        item_set = item_set | set(items)
    max_item = max(item_set)

    lines = open(sample_file).readlines()
    sample_seq = []
    for line in lines:
        user, items = line.strip().split(' ', 1)
        items = items.split(' ')
        items = [int(item) for item in items]
        sample_seq.append(items)

    assert len(user_seq) == len(sample_seq)

    return user_seq, max_item, sample_seq

def get_item2attribute_json(data_file):
    item2attribute = json.loads(open(data_file).readline())
    attribute_set = set()
    for item, attributes in item2attribute.items():
        attribute_set = attribute_set | set(attributes)
    attribute_size = max(attribute_set) # 331
    return item2attribute, attribute_size

def get_metric(pred_list, topk=10):
    NDCG = 0.0
    HIT = 0.0
    MRR = 0.0
    # [batch] the answer's rank
    for rank in pred_list:
        if rank < topk:
            MRR += 1.0 / (rank + 1.0)
            NDCG += 1.0 / np.log2(rank + 2.0)
            HIT += 1.0
    return HIT /len(pred_list), NDCG /len(pred_list), MRR /len(pred_list)

def precision_at_k_per_sample(actual, predicted, topk):
    num_hits = 0
    for place in predicted:
        if place in actual:
            num_hits += 1
    return num_hits / (topk + 0.0)

def precision_at_k(actual, predicted, topk):
    sum_precision = 0.0
    num_users = len(predicted)
    for i in range(num_users):
        act_set = set(actual[i])
        pred_set = set(predicted[i][:topk])
        sum_precision += len(act_set & pred_set) / float(topk)

    return sum_precision / num_users

def recall_at_k(actual, predicted, topk):
    sum_recall = 0.0
    num_users = len(predicted)
    true_users = 0
    for i in range(num_users):
        act_set = set(actual[i])
        pred_set = set(predicted[i][:topk])
        if len(act_set) != 0:
            sum_recall += len(act_set & pred_set) / float(len(act_set))
            true_users += 1
    return sum_recall / true_users


def apk(actual, predicted, k=10):
    """
    Computes the average precision at k.
    This function computes the average precision at k between two lists of
    items.
    Parameters
    ----------
    actual : list
             A list of elements that are to be predicted (order doesn't matter)
    predicted : list
                A list of predicted elements (order does matter)
    k : int, optional
        The maximum number of predicted elements
    Returns
    -------
    score : double
            The average precision at k over the input lists
    """
    if len(predicted)>k:
        predicted = predicted[:k]

    score = 0.0
    num_hits = 0.0

    for i,p in enumerate(predicted):
        if p in actual and p not in predicted[:i]:
            num_hits += 1.0
            score += num_hits / (i+1.0)

    if not actual:
        return 0.0

    return score / min(len(actual), k)


def mapk(actual, predicted, k=10):
    """
    Computes the mean average precision at k.
    This function computes the mean average prescision at k between two lists
    of lists of items.
    Parameters
    ----------
    actual : list
             A list of lists of elements that are to be predicted
             (order doesn't matter in the lists)
    predicted : list
                A list of lists of predicted elements
                (order matters in the lists)
    k : int, optional
        The maximum number of predicted elements
    Returns
    -------
    score : double
            The mean average precision at k over the input lists
    """
    return np.mean([apk(a, p, k) for a, p in zip(actual, predicted)])

def ndcg_k(actual, predicted, topk):
    res = 0
    for user_id in range(len(actual)):
        k = min(topk, len(actual[user_id]))
        idcg = idcg_k(k)
        dcg_k = sum([int(predicted[user_id][j] in
                         set(actual[user_id])) / math.log(j+2, 2) for j in range(topk)])
        res += dcg_k / idcg
    return res / float(len(actual))


# Calculates the ideal discounted cumulative gain at k
def idcg_k(k):
    res = sum([1.0/math.log(i+2, 2) for i in range(k)])
    if not res:
        return 1.0
    else:
        return res


def get_gpu_usage(device=None):
    r"""Return the reserved memory and total memory of given device in a string.
    Args:
        device: cuda.device. It is the device that the model run on.

    Returns:
        str: it contains the info about reserved memory and total memory of given device.
    """

    reserved = torch.cuda.max_memory_reserved(device) / 1024**3
    total = torch.cuda.get_device_properties(device).total_memory / 1024**3

    return "{:.2f} G/{:.2f} G".format(reserved, total)


def get_environment():
    device = "cuda"
    # print(device)
    gpu_usage = get_gpu_usage(device)

    import psutil

    memory_used = psutil.Process(os.getpid()).memory_info().rss / 1024**3
    memory_total = psutil.virtual_memory()[0] / 1024**3
    memory_usage = "{:.2f} G/{:.2f} G".format(memory_used, memory_total)
    cpu_usage = "{:.2f} %".format(psutil.cpu_percent(interval=1))
    """environment_data = [
        {"Environment": "CPU", "Usage": cpu_usage,},
        {"Environment": "GPU", "Usage": gpu_usage, },
        {"Environment": "Memory", "Usage": memory_usage, },
    ]"""

    table = Texttable()
    table.set_cols_align(["l", "c"])
    table.set_cols_valign(["m", "m"])
    table.add_rows(
        [
            ["Environment", "Usage"],
            ["CPU", cpu_usage],
            ["GPU", gpu_usage],
            ["Memory", memory_usage],
        ]
    )

    return table


def to_excel(result_info, args, training_time, inference_time, Epoch):
    # 转换result_info为字典，并将键中的下划线替换为@
    result_info = ast.literal_eval(result_info)
    result_info = {key.replace('_', '@'): value for key, value in result_info.items()}

    # 添加额外的信息
    result_info['Epoch'] = Epoch
    result_info['Training_time'] = training_time
    result_info['Inference_time'] = inference_time
    result_info['GPU'] = get_gpu_usage("cuda")
    result_info['DataSet'] = args.data_name
    result_info['Model'] = args.backbone
    keys_to_remove = ['MRR@5', 'MRR@10', 'MRR@20']
    for key in keys_to_remove:
        if key in result_info:
            del result_info[key]

    # 文件名
    file_name = 'Baseline_Full.xlsx'
    
    try:
        # 读取现有的Excel文件
        df = pd.read_excel(file_name)
    except FileNotFoundError:
        # 如果文件不存在，则创建一个新的DataFrame
        df = pd.DataFrame(columns=['DataSet', 'Model', 'HIT@5', 'HIT@10', 'HIT@20', 'NDCG@5', 'NDCG@10', 'NDCG@20', 'Training_time', 'GPU', 'Inference_time'])

    # 将新的结果转换为DataFrame
    new_result_df = pd.DataFrame([result_info])

    # 将新的结果添加到现有的DataFrame中
    df = pd.concat([df, new_result_df], ignore_index=True)
    # 按照'DataSet'和'Model'列进行排序
    df = df.sort_values(by=['DataSet', 'Model'])

    # 将更新后的DataFrame写入Excel文件
    df.to_excel(file_name, index=False)


    # 打开Excel文件
    workbook = openpyxl.load_workbook(file_name)

    # 遍历所有工作表
    for sheet in workbook.worksheets:
        # 遍历每一行
        for row in sheet.iter_rows():
            # 遍历每一个单元格
            for cell in row:
                # 设置单元格居中
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=False)

    # 保存修改后的文件
    workbook.save(file_name)

    # 加粗最大值
    workbook = load_workbook(file_name)
    sheet = workbook.active

    # 获取唯一的数据集值
    datasets = df['DataSet'].unique()

    # 定义需要处理的指标列
    metrics = ['HIT@5', 'HIT@10', 'HIT@20', 'NDCG@5', 'NDCG@10', 'NDCG@20']

    for dataset in datasets:
        dataset_df = df[df['DataSet'] == dataset]

        # 对每一个指标列
        for metric in metrics:
            if metric in dataset_df.columns:
                if dataset_df[metric].dtype == 'object':
                    dataset_df[metric] = pd.to_numeric(dataset_df[metric])

                # 找到最大值
                max_value = dataset_df[metric].max()

                # 遍历所有模型，并找出最大值所在的行
                for index, row in dataset_df.iterrows():
                    if row[metric] == max_value:
                        cell = sheet.cell(row=index + 2, column=df.columns.get_loc(metric) + 1)
                        cell.font = Font(bold=True)

    # 保存文件
    workbook.save(file_name)



def split_txt_file(filename, output_prefix, k, data_dir):
    # 读取文件内容
    with open(filename, 'r', encoding='utf-8') as file:
        content = file.readlines()

    # 计算每一份的行数
    total_lines = len(content)
    split_size = total_lines // k

    for i in range(1, k + 1)[::-1]:
        start_index = (i - 1) * split_size
        if i != k:
            end_index = i * split_size
        else:
            end_index = total_lines  # 最后一部分可能包括余下的所有行
        
        with open(f'{data_dir}/{output_prefix}_part_{k - i + 1}.txt', 'w', encoding='utf-8') as output_file:
            output_file.writelines(content[start_index:end_index])


def delete_files(output_prefix, parts, data_dir):
    for i in range(1, parts + 1):
        filename = str(data_dir) + f'{output_prefix}_part_{i}.txt'
        if os.path.exists(filename):
            os.remove(filename)
            print(f'文件 {filename} 已删除')
        else:
            print(f'文件 {filename} 不存在')